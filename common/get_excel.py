# -*- coding: utf-8 -*-
# @Time    : 2019/11/11 17:38
# @Author  : zhuxuefei

import openpyxl
import os
from config import Conf
from openpyxl import load_workbook

# 读取工作簿，表，以列表中元组的形式展示[('A2', 'B2', 'C2'), ('A3', 'B3', 'C3'), ('A4', 'B4', 'C4')]
def read_excel_tuple(bookname, sheetname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    if rows <= 1:
        print('无测试数据')
    else:
        param = []
        for i in range(1, rows):
            paramrow = list(ws.rows)[i]
            paramrow_value = []
            for j in range(0, columns):
                paramrow_value.append(paramrow[j].value)
            paramrow_value_tuple = tuple(paramrow_value)
            param.append(paramrow_value_tuple)
        return (param)

# 读取工作簿，表，列，该列以列表形式展示['A2', 'A3', 'A4', 'A5', 'A6']
def read_excel(bookname, sheetname, columnname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    for i in range(1, columns):
        nrows = list(ws.columns)[i]
        param = []
        if columnname == nrows[0].value:
            for j in range(1, rows):
                param.append(nrows[j].value)
            return (param)

if __name__=='__main__':
    testdata_path = Conf.get_testdata_path()
    excelFile = testdata_path + os.sep + "网络资源.xlsx"
    ## 创建VLAN池
    param_create_vlanpool = read_excel_tuple(excelFile, '创建VLAN池')
    print(param_create_vlanpool)
