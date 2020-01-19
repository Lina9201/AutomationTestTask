# -*- coding: utf-8 -*-
# @Time    : 2019/10/8 14:33
# @Author  : zhuxuefei

import xlrd
from openpyxl import load_workbook
import os

class OperationExcleData:
    """对excel进行操作，包括读取请求参数，和填写操作结果"""
    def __init__(self, excelFile, sheetName):
        # 验证文件是否存在，如果存在读取，不存在报错
        if os.path.exists(excelFile):
            self.excelFile = excelFile
            self.sheetName = sheetName
            self.caseList = []
        else:
            raise FileNotFoundError("文件不存在")

    def getCaseList(self):
        """
        获取excel中维护的测试用例，每个{}包含字段和对应值的键对值,一行测试用例也在dict中
        :param excelFile: 传入excel文件的名字
        :param sheetName: 传入接口所在sheet页的名字
        :return: 最终返回所有的测试用例
        """
        readExcel = xlrd.open_workbook(self.excelFile)
        sheet = readExcel.sheet_by_name(self.sheetName)
        # 获取excel的行数和列数
        trows = sheet.nrows
        tcols = sheet.ncols
        print(tcols)
        for i in range(2, trows):
            tmpdict = {}
            for j in range(2, tcols):
                case_key = sheet.cell_value(0, j)
                case_value = sheet.cell_value(i, j)
                if j < tcols - 1 and sheet.cell_value(0, j) != "" and sheet.cell_value(0, j + 1) == "":
                    childdict = {}
                if sheet.cell_value(0, j) != "" and sheet.cell_value(1, j) != "":
                    childdict = {}
                if sheet.cell_value(1, j) != "":
                    childcase_key = sheet.cell_value(1, j)
                    childdict[childcase_key] = sheet.cell_value(i, j)
                    case_value = childdict
                if case_key != "":
                    tmpdict[case_key] = case_value
            self.caseList.append(tmpdict)
        return self.caseList


    def getcase_tuple(self):
        wb = load_workbook(self.excelFile)
        ws = wb[self.sheetName]
        print(ws)
        # 获取excel的行数和列数
        trows = ws.max_row
        tcols = ws.max_column
        if trows <=1:
            print("没数据")
        else:
            for i in range(2, trows):
                case = list(ws.rows)[i]
                casevlaue = []
                for j in range(tcols):
                    casevlaue.append(case[j].value)
                casevlaue_tupe = tuple(casevlaue)
                self.caseList.append(casevlaue_tupe)
            print(self.caseList)
            return self.caseList


    def writeExcel(self,writeData):
        writeExcel = load_workbook(self.excelFile)
        try:
            writeSheet = writeExcel[self.sheetName]
            writeSheet["C2"]=writeData
            writeExcel.save(self.excelFile)
        except Exception as e:
            raise
        finally:
            pass











