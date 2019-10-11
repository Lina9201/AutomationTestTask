import pytest
import requests
from openpyxl import load_workbook

# 定义变量，以下函数都能调用
login_url_path = '/v1/tokens'


# 使用@pytest.fixture()，然后在同级别的文件中可以调用def函数
@pytest.fixture()
def ip():
    ip = '172.50.10.42'
    return ip


@pytest.fixture()
def port():
    port = '8000'
    return port


@pytest.fixture()
def login_json():
    login_json = {
        'authType': 'password',
        'params': {
            'username': 'duxiangyu',
            'password': 'eSXUb22UfzfFT+1L8/LinQ=='
        }
    }
    return login_json


@pytest.fixture()
def token(ip, port, login_json):
    ip_address = 'http://%s:%s' % (ip, port)
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json;charset=UTF-8'
    }
    login_reponse = requests.post(
        url=ip_address + login_url_path,
        json=login_json,
        headers=headers
    ).json()
    return login_reponse['data']['key']


@pytest.fixture()
def headers(token):
    headers = {
        'User-Agent': 'automation',
        'Content-Type': 'application/json;charset=UTF-8',
        'T-AUTH-TOKEN': token
    }
    return headers


# 读取工作簿，表，列，该列以列表形式展示['A2', 'A3', 'A4', 'A5', 'A6']
def read_excel(bookname, sheetname, columnname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    # print(list(ws.columns))
    for i in range(columns):
        a = list(ws.columns)[i]
        # print(a)
        b = []
        if columnname == a[0].value:
            for j in range(1, rows):
                b.append(a[j].value)
            return (b)


# 读取工作簿，表，以列表中元组的形式展示[('A2', 'B2', 'C2'), ('A3', 'B3', 'C3'), ('A4', 'B4', 'C4')]
def read_excel_tuple(bookname, sheetname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    if rows <= 1:
        print('没数据')
    else:
        param = []
        for i in range(1, rows):
            paramrow = list(ws.rows)[i]
            paramrow_value = []
            for j in range(columns):
                paramrow_value.append(paramrow[j].value)
            paramrow_value_tuple = tuple(paramrow_value)
            param.append(paramrow_value_tuple)
        return (param)


# 读取工作簿，表，以列表中字典的形式展示[{'A1': 'A2', 'B1': 'B2', 'C1': 'C2'}, {'A1': 'A3', 'B1': 'B3', 'C1': 'C3'}]
def read_excel_dic(bookname, sheetname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    title = list(ws.rows)[0]
    if rows <= 1:
        print('没数据')
    else:
        param = []
        key = []
        for i in range(len(title)):
            # 这是第一行数据，作为字典的key值
            key.append(title[i].value)
        print(key)
        j = 1
        for i in range(1, rows):
            paramrow = {}
            paramcol = list(ws.rows)[j]
            for x in range(columns):
                # 把key值对应的value赋值给key，每行循环
                paramrow[key[x]] = paramcol[x].value
            j += 1
            # 把字典加到列表中
            param.append(paramrow)
        return (param)


# 写入到工作簿，表，行，列，数据
def write_excel(bookname, sheetname, row, col, data):
    wb = load_workbook(bookname)  # 生成一个已存在的wookbook对象
    ws = wb.get_sheet_by_name(sheetname)  # 激活sheet
    ws.cell(row, col, data)  # 往sheet中的第二行第二列写入‘pass2’的数据
    wb.save(bookname)  # 保存
