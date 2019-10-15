# a = [1,2,3]
# b = ['a','b','c']
# d = []
# for i in range(len(a)):
#     a1 = (a[i],)
#     b1 = (b[i],)
#     c = a1 +b1
#     print(c)
#     d.append(c)
# print(d)
from openpyxl import load_workbook
def read_excel_tuple(bookname, sheetname):
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    if rows <= 1:
        print('没数据')
    else:
        param = []
        for i in range(1,rows):
            paramrow = list(ws.rows)[i]
            paramrow_value = []
            for j in range(columns):
                paramrow_value.append(paramrow[j].value)
            paramrow_value_tuple = tuple(paramrow_value)
            param.append(paramrow_value_tuple)
        return (param)
print(read_excel_tuple('测试数据.xlsx', '创建VLAN池'))

str = '[{"resourcePoolId": 114}]'
a = str.split(',')
print(a)
print(type(a))
b = a[0]
print(b)
print(type(b))