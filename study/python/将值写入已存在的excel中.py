# # from xlutils.copy import copy
# # import xlrd
# # import xlwt
# # # from xlrd import open_workbook
# # from xlwt import easyfont
# # import os
# #
# # def writeexcel():
# #     excel = r'D:\测试数据.xls'
# #     # 打开文件
# #     openworkbook = xlrd.open_workbook(excel,formatting_info = True)
# #     # 复制文件
# #     copyworkbook =copy(openworkbook)
# #     # 打开表格
# #     table = openworkbook.sheet_by_name("编辑网络")
# #     # 写入数据
# #     table.write(2,1,"147")
# #     os.remove(excel)
# #     copyworkbook.save(excel)
# # writeexcel()
# from xlutils.copy import copy
# import xlrd
#
# # 加载已存在的xls
# old_workbook = xlrd.open_workbook('data.xls')
#
# # 将已存在的excel拷贝进新的excel
# new_workbook = copy(old_workbook)
#
# # 获取sheet
# new_worksheet = new_workbook.get_sheet(0)
#
# # 写入数据
# row = 5  # 已存在文件中的数据行数
# for data in data_set:
#     new_worksheet.write(row, 0, data['href'])
# new_worksheet.write(row, 1, data['star'])
# row += 1
#
# # 将新写入的数据保存
# new_workbook.save('data.xls')
# coding:utf-8
import openpyxl


# 写入已存在的xlsx文件第一种方法
# class Write_excel(object):
#     '''修改excel数据'''
#     def __init__(self, filename):
#         self.filename = filename
#         self.wb = load_workbook(self.filename)
#         self.ws = self.wb.active  # 激活sheet
#
#     def write(self, row_n, col_n, value):
#         '''写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"'''
#         self.ws.cell(row_n, col_n,value )
#         self.wb.save(self.filename)
#
# we = Write_excel("mylogintest.xlsx")
# we.write(2,2,'pass3')


import  openpyxl
#写入已存在的xlsx文件第二种方法
# from openpyxl import load_workbook
# wb = load_workbook("测试数据.xlsx")#生成一个已存在的wookbook对象
# ws = wb["Sheet1"] # 激活sheet
# ws.cell(2,1,'1')#往sheet中的第二行第二列写入‘pass2’的数据
# wb.save("测试数据.xlsx")#保存

from openpyxl import load_workbook
# 打开已存在的workbook
wb = load_workbook("测试数据.xlsx")
# 通过sheet名打开sheet页
ws = wb['Sheet1']
# 获取sheet页最大行数，最大列数
rows = ws.max_row
columns = ws.max_column
print(rows,columns)
# # 按照行，列出所有的单元格
# a = list(ws.rows)[0]
# print(a)
# b = []
# for i in range(len(a)):
#     b.append(a[i].value)
# print(b)
# # 按照列，列出所有的单元格
# a = list(ws.columns)
# print(a)
title = list(ws.rows)[0]
if rows <= 1:
    print("没数据")
else:
    param = []
    key = []
    for i in range(len(title)):
        key.append(title[i].value)  # 这是第一行数据，作为字典的key值
    print(key)
    j = 1
    for i in range(1,rows):
        paramrow ={}
        paramcol = list(ws.rows)[j]
        for x in range(columns):
            # 把key值对应的value赋值给key，每行循环
            paramrow[key[x]]=paramcol[x].value
        j+=1
        # 把字典加到列表中
        param.append(paramrow)
        print(param)