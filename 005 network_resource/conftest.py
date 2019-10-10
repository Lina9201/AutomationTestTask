import pytest
import requests
import xlrd
import xlwt
from openpyxl import load_workbook

#定义变量，以下函数都能调用
login_url_path = '/v1/tokens'

#使用@pytest.fixture()，然后在同级别的文件中可以调用def函数
@pytest.fixture()
def ip():
    ip = '172.50.10.42'
    return ip

@pytest.fixture()
def port():
    port = '8000'
    return port

@pytest.fixture()
def login_json():
    login_json = {'authType': "password",
                'params': {'username': "duxiangyu", 'password': "eSXUb22UfzfFT+1L8/LinQ=="}
                }
    return login_json

@pytest.fixture()
def token(ip,port,login_json):
    ip_address = 'http://%s:%s'%(ip,port)
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json;charset=UTF-8',
   }

    login_request = requests.post(url=ip_address+login_url_path,json=login_json,headers=headers)
    login_reponse = login_request.json()
    return login_reponse['data']['key']

@pytest.fixture()
def headers(token):
    headers = {
        "User-Agent": "automation",
        'Content-Type': 'application/json;charset=UTF-8',
        'T-AUTH-TOKEN': token
        }
    return headers

def readexcel(dataname,Sheetname):
    #打开excel
    workboll = xlrd.open_workbook(r'%s' % dataname)
    # workboll = xlrd.open_workbook(r'D:\测试数据.xlsx')

    #打印页签名
    print(workboll.sheet_names())

    #打开页签
    # sheetname = workboll.sheet_by_name("Sheet1")
    sheetname = workboll.sheet_by_name(Sheetname)
    #获取行数
    nrows = sheetname.nrows

    #获取列数
    ncols = sheetname.ncols

    #打印行数和列数
    print(nrows,ncols)

    sheetvlaues = []
    sheetvlauesa = []

    for i in range(1,nrows):
        for j in range(ncols):
            sheetvlauesa.append(sheetname.cell(i, j).value)
        sheetvlaues.append(tuple(sheetvlauesa))
        sheetvlauesa = []
    return sheetvlaues
    # print(sheetvlaues)
# print(readexcel("D:\测试数据1.xlsx","B"))


def read_excel_dic(bookname,sheetname):
    # #打开excel表，填写路径
    # openworkbook = xlrd.open_workbook(r'%s' % bookname)
    # #找到sheet页
    # table = openworkbook.sheet_by_name(sheetname)
    # #获取总行数总列数
    # row_Num = table.nrows
    # col_Num = table.ncols
    #
    # param =[]
    # key =table.row_values(0)# 这是第一行数据，作为字典的key值
    #
    # if row_Num <= 1:
    #     print("没数据")
    # else:
    #     j = 1
    #     for i in range(1,row_Num):
    #         paramrow ={}
    #         paramcol = table.row_values(j)
    #         for x in range(col_Num):
    #             # 把key值对应的value赋值给key，每行循环
    #             paramrow[key[x]]=paramcol[x]
    #         j+=1
    #         # 把字典加到列表中
    #         param.append(paramrow)
    #     return param
    # 打开已存在的workbook
    wb = load_workbook(bookname)
    # 通过sheet名打开sheet页
    ws = wb[sheetname]
    # 获取sheet页最大行数，最大列数
    rows = ws.max_row
    columns = ws.max_column
    title = list(ws.rows)[0]
    if rows <= 1:
        print("没数据")
    else:
        param = []
        key = []
        for i in range(len(title)):
            key.append(title[i].value)  # 这是第一行数据，作为字典的key值
        print(key)
        j = 1
        for i in range(1,rows):
            paramrow ={}
            paramcol = list(ws.rows)[j]
            for x in range(columns):
                # 把key值对应的value赋值给key，每行循环
                paramrow[key[x]]=paramcol[x].value
            j+=1
            # 把字典加到列表中
            param.append(paramrow)
        return (param)

def write_excel(bookname,sheetname,row,col,data):
    wb = load_workbook(bookname)#生成一个已存在的wookbook对象
    ws = wb.get_sheet_by_name(sheetname)#激活sheet
    ws.cell(row,col,data)#往sheet中的第二行第二列写入‘pass2’的数据
    wb.save(bookname)#保存